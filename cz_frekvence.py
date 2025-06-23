import requests
import json
import csv
from openpyxl import load_workbook
from bs4 import BeautifulSoup

frequency_file = 'CZ_frekvence_2024.xlsx'  # Read aerodrome codes to retrieve freqencies, and airport names and frequencies for json generation


def get_frequency(aerodrome):
    """Retrieves frequency for aerodrome ICAO code from aim.rlp.cz website"""
    response = requests.get(f'https://aim.rlp.cz/vfrmanual/actual/{str(aerodrome).lower()}_text_cz.html')
    soup = BeautifulSoup(response.text, 'html.parser')

    # '<div id="aerodrome-frekvence"><img src="ad/icons/frekvence.png" class="aerodrome-icon" alt="Frekvence">'
    # dec = -1
    res = ''
    content = soup.find('div', {'id': 'aerodrome-frekvence'})
    pos = 0

    if content is None:
        return f'Not found: {aerodrome}'

    for i in content.text:
        if i.isdigit() or i == ',':
            space = ''
            # if dec > 2:
            #     space = ''

            if pos == 7:
                space = ' '
                pos = 0

            plus = i

            if i == ',':
                plus = '.'

            res = res + (space + plus)
            pos += 1

            # if i == ',':
            #     dec = 0
            # if i.isdigit() and -1 < dec < 3:
            #     dec += 1

    if ' ' in res:
        return content.text
    print(f'{aerodrome}: {res}')
    print()
    return res
    # ''.join([i for i in soup.find('div', {'id': 'aerodrome-frekvence'}).text if i.isdigit() or i == ','])


def retrieve_frequencies(frequency_file):
    """
    Read airport ICAO code from column A of Excel file.
    Extract airport frequency from aim.rlp.cz website.
    Write frequency to column B of Excel file.
    Manual checking of retrieved frequencies required.
    """
    print(frequency_file)

    wb = load_workbook(frequency_file)
    ws = wb.active
    max_row = ws.max_row

    print(f'file: {frequency_file}, sheet {ws}, entries: {max_row}')

    for i in range(1, max_row + 1):
        cell = ws['A' + str(i)].value   # read airport ICAO code from column A
        if cell:
            values = cell.split(',')
            if len(values) > 1:
                icao_code = values[1].strip('”')
            else:
                icao_code = cell
            print(f'{i} - Retrieving frequency of {icao_code}')
            airport_frequency = get_frequency(icao_code)  # get airport frequency
            ws['B' + str(i)] = airport_frequency  # write airport frequency to column B

    wb.save(frequency_file)


def create_json_file(frequency_file):
    """
    Read airport name and frequency from Excel file.
    Create a json file to be loaded into KRT2Bluetooth Android app.
    """
    json_output_file = frequency_file.strip('.xlsx') + '.json'
    json_data = []

    wb = load_workbook(frequency_file)
    ws = wb.active
    max_row = ws.max_row

    print(f'file: {frequency_file}, sheet {ws}, entries: {max_row},')

    try:
        with open(json_output_file, 'w') as file:

            for i in range(1, max_row + 1):
                airport = ws['C' + str(i)].value  # read airport name from column C
                frequency = ws['B' + str(i)].value  # read airport frequency from column B
                # frequency = "{:.3f}".format(float(frequency))  # add trailing zeroes to frequency

                print(airport, frequency)

                json_data.append({"idx": i - 1, "label": airport, "freq": float(frequency)})

            all_json_data = {"name": frequency_file.replace('_frekvence', '_freq').strip('.xlsx'), "items": json_data}
            file.write(json.dumps(all_json_data))

        print("Data added to JSON file successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")


def create_csv_file(frequency_file):
    """
    Read airport name and frequency from Excel file.
    Create a csv file to be loaded into KRT2 Manager PC app.
    """
    csv_output_file = frequency_file.strip('.xlsx') + '.csv'

    wb = load_workbook(frequency_file)
    ws = wb.active
    max_row = ws.max_row

    print(f'file: {frequency_file}, sheet {ws}, entries: {max_row},')

    try:
        with open(csv_output_file, 'w', newline='') as file:
            writer = csv.writer(file)
            for i in range(1, max_row + 1):
                airport = ws['C' + str(i)].value  # read airport name from column C
                frequency = ws['B' + str(i)].value  # read airport frequency from column B
                # frequency = "{:.3f}".format(float(frequency))  # add trailing zeroes to frequency

                csv_data = [airport, frequency]
                print(csv_data)

                writer.writerow(csv_data)
        print("Data added to CSV file successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    retrieve_frequencies(frequency_file)  # Read aerodrome ICAO codes from Excel file and retrieve frequencies from web, write to the same file
    create_json_file(frequency_file)  # Read airport names and frequencies from frequency Excel file, generate new json file
    create_csv_file(frequency_file)  # Read airport names and frequencies from frequency Excel file, generate new csv file

