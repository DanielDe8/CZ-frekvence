import requests
from openpyxl import *
from bs4 import BeautifulSoup


def main():
    # file = open("sequence.xlsx")
    # ws = file.worksheets[0]
    # ws['A2'] = 'ahoj z pajtnu'
    # file.()

    file_path = input('Zadejte soubor: ')  # 'sequence.xlsx'

    wb = load_workbook(file_path)

    ws = wb.active  # or wb.active

    max_row = ws.max_row

    for i in range(1, max_row + 1):
        cell = ws['A' + str(i)].value
        values = cell.split(',')
        if len(values) > 1:
            res = values[1].strip('”')
        else:
            res = cell
        ws['B' + str(i)] = frequence(res)

    wb.save(file_path)

    # wb = Workbook()
    # ws = wb.active
    # wb.save(input("Zadejte jmeno souboru: "))


def frequence(letiste):
    req = requests.get(f'https://aim.rlp.cz/vfrmanual/actual/{str(letiste).lower()}_text_cz.html')
    soup = BeautifulSoup(req.text, 'html.parser')
    # '<div id="aerodrome-frekvence"><img src="ad/icons/frekvence.png" class="aerodrome-icon" alt="Frekvence">'
    # dec = -1
    res = ''
    content = soup.find('div', {'id': 'aerodrome-frekvence'})
    pos = 0

    if content is None:
        return f'Not found: {letiste}'

    for i in content.text:
        if i.isdigit() or i == ',':
            space = ''
            # if dec > 2:
            #     space = ''

            if pos == 7:
                space = ' '
                pos = 0

            plus = i

            if i == ',':
                plus = '.'

            res = res + (space + plus)
            pos += 1

            # if i == ',':
            #     dec = 0
            # if i.isdigit() and -1 < dec < 3:
            #     dec += 1

    if ' ' in res:
        return content.text
    return res
    # ''.join([i for i in soup.find('div', {'id': 'aerodrome-frekvence'}).text if i.isdigit() or i == ','])


if __name__ == "__main__":
    main()
